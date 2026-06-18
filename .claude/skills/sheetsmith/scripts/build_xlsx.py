#!/usr/bin/env python3
"""build_xlsx.py — build a styled .xlsx workbook (incl. themed dashboards) from a JSON spec.

Usage:
    python3 build_xlsx.py <spec.json> <out.xlsx>

Used by the `sheetsmith` skill. The skill drafts the workbook spec, adversarial
sub-agents audit it, then this script materializes it into a styled .xlsx —
including dashboard tabs with KPI cards, palette-themed charts, section bars,
spacer-grid column widths, range fills/borders, and rich conditional formats.
See ../references/spec-schema.md for the spec format and ../references/dashboard-layout.md
for the layout/color system.

Requires: openpyxl  (pip install openpyxl)
"""
import json
import re
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.formatting.rule import (
        CellIsRule,
        ColorScaleRule,
        DataBarRule,
        FormulaRule,
        IconSetRule,
    )
    from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
except ImportError:
    sys.exit("build_xlsx.py needs openpyxl — install it with:  pip install openpyxl")


# Default theme — the "Budget by Paycheck" palette: soft pastels + a terracotta accent.
# Charts pull series/slice colors from chart_palette so nothing falls back to Excel defaults.
DEFAULT_THEME = {
    "accent": "EAC6B8",
    "header_font": "FFFFFF",
    "border": "E8DAD3",
    "section_colors": ["F6E5A8", "F4C0DB", "F8B6C2", "99E4EB", "ABA5E3"],
    "body_tints": ["ECF3F1", "FAF1F3", "FDFBF3", "E6F3F4", "EDE8F8"],
    "chart_palette": ["7ED2B6", "F8B6C2", "80D2DA", "B9A4E7", "F6E5A8", "F4C0DB", "A3CEC5", "EAE9F4"],
}
CHART_TYPES = {"bar": BarChart, "line": LineChart, "pie": PieChart, "doughnut": DoughnutChart}


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _border(color, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _as_list(value):
    """openpyxl rule `formula` is a list of strings; accept a bare string too."""
    return value if isinstance(value, list) else [value]


def _col_letter_of(cell_ref):
    m = re.match(r"([A-Za-z]+)", cell_ref)
    return m.group(1).upper() if m else None


def _vertical_count(ref_str):
    """How many cells a range spans (max of row/col span)."""
    minc, minr, maxc, maxr = range_boundaries(ref_str.split("!")[-1])
    return max(maxr - minr + 1, maxc - minc + 1)


def _reference(wb, default_ws, ref_str):
    """Build an openpyxl Reference from 'Tab!A1:A10' or 'A1:A10' (default tab)."""
    if "!" in ref_str:
        sheet_name, rng = ref_str.split("!", 1)
        ws = wb[sheet_name]
    else:
        ws, rng = default_ws, ref_str
    min_col, min_row, max_col, max_row = range_boundaries(rng)
    return Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)


# ---------- dashboard components ----------

def _add_kpi_card(ws, cfg, theme):
    """A merged, styled stat tile: a small label over a big number/formula."""
    r, c = coordinate_to_tuple(cfg["anchor"])  # (row, col)
    rows, cols = cfg.get("span", [2, 2])
    rows = max(rows, 2)  # need a label row + at least one value row
    fill = _fill(cfg.get("fill", theme["accent"]))
    font_color = cfg.get("font_color", theme["header_font"])
    border = _border(theme["border"]) if cfg.get("border") else None
    # style the whole card BEFORE merging (merged non-anchor cells are read-only)
    for rr in range(r, r + rows):
        for cc in range(c, c + cols):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = fill
            if border:
                cell.border = border
    lab = ws.cell(row=r, column=c, value=cfg.get("label", ""))
    lab.font = Font(bold=True, size=cfg.get("label_size", 10), color=font_color)
    lab.alignment = Alignment(horizontal="center", vertical="center")
    val = ws.cell(row=r + 1, column=c, value=cfg.get("value"))
    val.font = Font(bold=True, size=cfg.get("value_size", 20), color=font_color)
    val.alignment = Alignment(horizontal="center", vertical="center")
    if cfg.get("number_format"):
        val.number_format = cfg["number_format"]
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + cols - 1)
    ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + rows - 1, end_column=c + cols - 1)


def _add_section_bar(ws, cfg, theme, idx):
    """A merged, colored header bar that titles a panel (no chart titles — these do the labeling)."""
    rng = cfg["range"]
    color = cfg.get("color") or theme["section_colors"][idx % len(theme["section_colors"])]
    minc, minr, _, _ = range_boundaries(rng.split("!")[-1])
    for row in ws[rng]:
        for cell in row:
            cell.fill = _fill(color)
    ws.merge_cells(rng)
    anchor = ws.cell(row=minr, column=minc, value=cfg.get("label", ""))
    anchor.font = Font(bold=True, size=cfg.get("size", 11), color=cfg.get("font_color", theme["header_font"]))
    anchor.alignment = Alignment(horizontal=cfg.get("align", "center"), vertical="center")


def _theme_chart(chart, kind, cfg, n_points, palette):
    """Color series (bar/line) or individual slices (pie/doughnut) from the theme palette."""
    colors = cfg.get("colors") or palette
    if kind in ("bar", "line"):
        for i, s in enumerate(chart.series):
            s.graphicalProperties = GraphicalProperties(solidFill=colors[i % len(colors)])
    elif chart.series:  # pie / doughnut → color each slice
        ser = chart.series[0]
        for i in range(max(n_points, 0)):
            dp = DataPoint(idx=i)
            dp.graphicalProperties = GraphicalProperties(solidFill=colors[i % len(colors)])
            ser.data_points.append(dp)


def _add_chart(wb, ws, cfg, theme):
    """Bar/line/pie/doughnut chart anchored on `ws`, data optionally drawn from another tab."""
    kind = cfg.get("type", "bar")
    chart = CHART_TYPES[kind]()
    if cfg.get("title"):
        chart.title = cfg["title"]
    tfd = cfg.get("titles_from_data", kind in ("bar", "line"))
    chart.add_data(_reference(wb, ws, cfg["data"]), titles_from_data=tfd)
    if cfg.get("categories"):
        chart.set_categories(_reference(wb, ws, cfg["categories"]))
        n_points = _vertical_count(cfg["categories"])
    else:
        n_points = _vertical_count(cfg["data"]) - (1 if tfd else 0)
    _theme_chart(chart, kind, cfg, n_points, theme["chart_palette"])
    chart.height = cfg.get("height", 7)
    chart.width = cfg.get("width", 12)
    ws.add_chart(chart, cfg.get("anchor", "A1"))


def _add_conditional_format(ws, cf):
    ctype = cf.get("type", "cell_is")
    fill = _fill(cf.get("fill", "FFC7CE"))
    if ctype == "formula":
        ws.conditional_formatting.add(cf["range"], FormulaRule(formula=_as_list(cf["formula"]), fill=fill))
    elif ctype == "data_bar":
        ws.conditional_formatting.add(
            cf["range"], DataBarRule(start_type="min", end_type="max", color=cf.get("color", "638EC6"))
        )
    elif ctype == "color_scale":
        ws.conditional_formatting.add(
            cf["range"],
            ColorScaleRule(
                start_type="min", start_color=cf.get("min_color", "F8696B"),
                mid_type="percentile", mid_value=50, mid_color=cf.get("mid_color", "FFEB84"),
                end_type="max", end_color=cf.get("max_color", "63BE7B"),
            ),
        )
    elif ctype == "icon_set":
        ws.conditional_formatting.add(
            cf["range"],
            IconSetRule(
                icon_style=cf.get("icon_style", "3TrafficLights1"),
                type="percent",
                values=cf.get("values", [0, 33, 67]),
                reverse=cf.get("reverse", False),
            ),
        )
    else:  # cell_is
        ws.conditional_formatting.add(
            cf["range"],
            CellIsRule(
                operator=cf.get("operator", "lessThan"),
                formula=[str(x) for x in _as_list(cf.get("formula", ["0"]))],
                fill=fill,
            ),
        )


def validate(spec):
    """Fail loud and specific on a malformed spec rather than silently building garbage."""
    if not isinstance(spec, dict):
        raise ValueError("spec must be a JSON object")
    tabs = spec.get("tabs")
    if not tabs:
        raise ValueError("spec.tabs is missing or empty")
    for i, tab in enumerate(tabs):
        name = tab.get("name")
        if not name:
            raise ValueError(f"tab[{i}] is missing 'name'")
        # a dashboard tab may have no columns if it carries cards/charts/section bars
        if not tab.get("columns") and not (tab.get("kpi_cards") or tab.get("charts") or tab.get("section_bars")):
            raise ValueError(f"tab '{name}' has no 'columns' (and no kpi_cards/charts/section_bars either)")
        for f in tab.get("formulas", []):
            if "range_rows" in f:
                rr = f.get("range_rows")
                if not (isinstance(rr, list) and len(rr) == 2 and rr[0] <= rr[1]):
                    raise ValueError(
                        f"tab '{name}': formula range_rows must be [start, end] with start<=end, got {rr!r}"
                    )
                if "col" not in f:
                    raise ValueError(f"tab '{name}': templated formula with range_rows is missing 'col'")
            elif "cell" not in f:
                raise ValueError(f"tab '{name}': each formula needs 'cell' or ('col' + 'range_rows')")
        for ch in tab.get("charts", []):
            if "data" not in ch:
                raise ValueError(f"tab '{name}': a chart is missing its 'data' range")
            if ch.get("type", "bar") not in CHART_TYPES:
                raise ValueError(f"tab '{name}': chart type must be {sorted(CHART_TYPES)}, got {ch.get('type')!r}")
        for card in tab.get("kpi_cards", []):
            if "anchor" not in card:
                raise ValueError(f"tab '{name}': a kpi_card is missing its 'anchor' cell")
        for key in ("fills", "borders", "section_bars"):
            for item in tab.get(key, []):
                if "range" not in item:
                    raise ValueError(f"tab '{name}': a '{key}' entry is missing its 'range'")


def build(spec, out_path):
    """Build the workbook described by `spec` and save to out_path. Returns a summary list."""
    validate(spec)
    theme = {**DEFAULT_THEME, **spec.get("theme", {})}
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    summary = []

    for tab in spec["tabs"]:
        ws = wb.create_sheet(title=tab["name"][:31])  # Excel caps tab names at 31 chars
        cols = tab.get("columns", [])
        styles = tab.get("styles", {})
        header_style = styles.get("header", {})
        col_fmt = {}  # column-letter -> number_format, so formula cells can inherit it

        # --- header row ---
        for c, col in enumerate(cols, start=1):
            letter = get_column_letter(c)
            if col.get("number_format"):
                col_fmt[letter] = col["number_format"]
            cell = ws.cell(row=1, column=c, value=col.get("header", ""))
            cell.font = Font(bold=header_style.get("bold", True), color=header_style.get("font_color", "FFFFFF"))
            if header_style.get("fill"):
                cell.fill = _fill(header_style["fill"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col.get("width"):
                ws.column_dimensions[letter].width = col["width"]

        # --- data rows ---
        rows = tab.get("rows", [])
        for r, rowdata in enumerate(rows, start=2):
            for c, col in enumerate(cols, start=1):
                key = col.get("key")
                if key is not None and key in rowdata:
                    cell = ws.cell(row=r, column=c, value=rowdata[key])
                    if col.get("number_format"):
                        cell.number_format = col["number_format"]

        # --- row banding (even data rows) ---
        banding = styles.get("banding", {})
        if banding.get("enabled"):
            band_fill = _fill(banding.get("color", "F2F2F2"))
            for r in range(2, 2 + len(rows)):
                if r % 2 == 0:
                    for c in range(1, len(cols) + 1):
                        ws.cell(row=r, column=c).fill = band_fill

        # --- formulas (a formula cell inherits its column's number_format) ---
        for f in tab.get("formulas", []):
            if "cell" in f:
                ref = f["cell"]
                ws[ref] = f["formula"]
                fmt = f.get("number_format") or col_fmt.get(_col_letter_of(ref))
                if fmt:
                    ws[ref].number_format = fmt
            else:  # col + range_rows (validated above)
                start, end = f["range_rows"]
                col_letter = f["col"].upper()
                fmt = f.get("number_format") or col_fmt.get(col_letter)
                for row in range(start, end + 1):
                    ref = f"{col_letter}{row}"
                    ws[ref] = f["formula"].replace("{row}", str(row))
                    if fmt:
                        ws[ref].number_format = fmt

        # --- conditional formats ---
        for cf in tab.get("conditional_formats", []):
            _add_conditional_format(ws, cf)

        # --- layout: spacer-grid column widths + uniform row height ---
        for letter, w in tab.get("column_widths", {}).items():
            ws.column_dimensions[letter.upper()].width = w
        if tab.get("row_height"):
            ws.sheet_format.defaultRowHeight = tab["row_height"]
            ws.sheet_format.customHeight = True

        # --- range fills + borders (apply BEFORE merges to avoid read-only MergedCell) ---
        for fcfg in tab.get("fills", []):
            fill = _fill(fcfg["color"])
            for row in ws[fcfg["range"]]:
                for cell in row:
                    cell.fill = fill
        for bcfg in tab.get("borders", []):
            bd = _border(bcfg.get("color", theme["border"]), bcfg.get("style", "thin"))
            for row in ws[bcfg["range"]]:
                for cell in row:
                    cell.border = bd

        # --- dashboard components ---
        for idx, bar in enumerate(tab.get("section_bars", [])):
            _add_section_bar(ws, bar, theme, idx)
        for card in tab.get("kpi_cards", []):
            _add_kpi_card(ws, card, theme)
        for chart_cfg in tab.get("charts", []):
            _add_chart(wb, ws, chart_cfg, theme)

        # --- view options ---
        if tab.get("freeze"):
            ws.freeze_panes = tab["freeze"]
        if tab.get("hide_gridlines"):
            ws.sheet_view.showGridLines = False

        summary.append(
            f"{tab['name']}: {len(cols)} cols, {len(rows)} rows, "
            f"{len(tab.get('formulas', []))} formula block(s), "
            f"{len(tab.get('conditional_formats', []))} cond-format(s), "
            f"{len(tab.get('section_bars', []))} bar(s), "
            f"{len(tab.get('kpi_cards', []))} KPI card(s), {len(tab.get('charts', []))} chart(s)"
        )

    wb.save(out_path)
    return summary


def main():
    if len(sys.argv) != 3:
        sys.exit("usage: python3 build_xlsx.py <spec.json> <out.xlsx>")
    spec_path, out_path = sys.argv[1], sys.argv[2]
    with open(spec_path) as fh:
        spec = json.load(fh)
    summary = build(spec, out_path)
    print(f"Built {out_path}", file=sys.stderr)
    for line in summary:
        print("  " + line, file=sys.stderr)


if __name__ == "__main__":
    main()
